"""Excel 報告：一個解析度一個頁籤 + 總覽頁；BUG 步驟內嵌 原影片/點擊前/點擊後 縮圖。

主要訴求：找各解析度下、模擬器跑是否有「圖歪 / 掉圖」等適配問題，
並標出「點擊後畫面跑掉」的 BUG，計算每個解析度的通過率。
"""
from __future__ import annotations

from pathlib import Path

import cv2

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:  # pragma: no cover
    raise ImportError("需要 openpyxl：請執行 py -m pip install openpyxl") from e

from .runner import Suite

THUMB_W = 220
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)


def _thumb(src: Path, dst_dir: Path, name: str) -> Path | None:
    img = cv2.imread(str(src))
    if img is None:
        return None
    h, w = img.shape[:2]
    nw = THUMB_W
    nh = max(1, int(h * nw / w))
    small = cv2.resize(img, (nw, nh), interpolation=cv2.INTER_AREA)
    dst_dir.mkdir(parents=True, exist_ok=True)
    out = dst_dir / name
    cv2.imwrite(str(out), small)
    return out


def write_excel(suite: Suite, root: Path) -> Path:
    thumbs_dir = root / "thumbs"
    wb = Workbook()
    _build_summary(wb.active, suite)
    for res in suite.resolutions:
        ws = wb.create_sheet(_safe_sheet(res))
        _build_resolution_sheet(ws, suite, res, root, thumbs_dir)
    out = root / "report.xlsx"
    wb.save(out)
    return out


def _safe_sheet(name: str) -> str:
    # Excel 頁籤名不可含 []:*?/\ 且 <=31 字
    for ch in r"[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:31]


def _build_summary(ws, suite: Suite):
    ws.title = "總覽"
    ws["A1"] = "自動化測試報告（適配檢查）"
    ws["A1"].font = TITLE
    ws["A2"] = f"腳本：{suite.script_name}"
    ws["A3"] = f"時間：{suite.started_at}　每解析度重複：{suite.repeat} 次"
    ws["A5"] = f"總成功率：{suite.success_rate():.1f}%"
    ws["A5"].font = BOLD

    headers = ["解析度", "通過/總數", "成功率", "BUG 步驟數", "崩潰/ANR 次數"]
    r = 7
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = BOLD
        cell.fill = HEAD_FILL
    for res in suite.resolutions:
        runs = [x for x in suite.runs if x.resolution == res]
        passed = sum(1 for x in runs if x.passed)
        bugs = sum(x.bug_count for x in runs)
        crashes = sum(1 for x in runs if x.crashes)
        r += 1
        ws.cell(r, 1, res)
        ws.cell(r, 2, f"{passed}/{len(runs)}")
        rate = suite.success_rate(res)
        rc = ws.cell(r, 3, f"{rate:.1f}%")
        rc.fill = PASS_FILL if rate >= 80 else FAIL_FILL
        ws.cell(r, 4, bugs)
        ws.cell(r, 5, crashes)
    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 16


def _build_resolution_sheet(ws, suite: Suite, res: str, root: Path, thumbs_dir: Path):
    runs = [x for x in suite.runs if x.resolution == res]
    passed = sum(1 for x in runs if x.passed)
    rate = suite.success_rate(res)

    ws["A1"] = f"解析度 {res}"
    ws["A1"].font = TITLE
    ws["A2"] = f"測試次數：{len(runs)}　通過：{passed}　成功率：{rate:.1f}%"
    ws["A2"].font = BOLD

    # 每輪摘要表
    r = 4
    for c, h in enumerate(["第幾次", "結果", "耗時(s)", "BUG步驟", "崩潰/ANR", "錯誤訊息"], 1):
        cell = ws.cell(r, c, h)
        cell.font = BOLD
        cell.fill = HEAD_FILL
    for run in runs:
        r += 1
        ws.cell(r, 1, run.run_index)
        rc = ws.cell(r, 2, "PASS" if run.passed else "FAIL")
        rc.fill = PASS_FILL if run.passed else FAIL_FILL
        ws.cell(r, 3, run.duration_sec)
        ws.cell(r, 4, run.bug_count)
        ws.cell(r, 5, len(run.crashes))
        ws.cell(r, 6, (run.error or "").split("\n")[0][:80])

    # BUG 明細（內嵌縮圖）
    r += 2
    ws.cell(r, 1, "BUG 明細（圖歪 / 掉圖 / 點擊無反應 / 點擊後跑掉）").font = TITLE
    r += 1
    for c, h in enumerate(["第幾次/步驟", "原因", "相似度(前/後)",
                           "原影片", "點擊前", "點擊後/結果"], 1):
        cell = ws.cell(r, c, h)
        cell.font = BOLD
        cell.fill = HEAD_FILL
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    for col in ("D", "E", "F"):
        ws.column_dimensions[col].width = 34

    any_bug = False
    for run in runs:
        sdir = Path(run.screenshot_dir)
        for s in run.bug_steps:
            any_bug = True
            r += 1
            ws.row_dimensions[r].height = 110
            ws.cell(r, 1, f"#{run.run_index} / {s.name}").alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, 2, s.bug_reason).alignment = Alignment(wrap_text=True, vertical="top")
            sim = (f"{s.ref_similarity:.2f}" if s.ref_similarity is not None else "—") + \
                  " / " + (f"{s.after_similarity:.2f}" if s.after_similarity is not None else "—")
            ws.cell(r, 3, sim).alignment = Alignment(vertical="top")
            # 三張縮圖
            for col_letter, fname in (("D", s.ref_shot), ("E", s.before_shot),
                                      ("F", s.screenshot)):
                if not fname:
                    continue
                t = _thumb(sdir / fname, thumbs_dir, f"{res}_{run.run_index}_{s.index}_{col_letter}.png")
                if t:
                    img = XLImage(str(t))
                    ws.add_image(img, f"{col_letter}{r}")

        # 崩潰摘要附在最後
        if run.crashes:
            any_bug = True
            r += 1
            ws.cell(r, 1, f"#{run.run_index} 崩潰/ANR")
            ws.cell(r, 2, " | ".join(run.crashes[:3])[:300]).alignment = Alignment(wrap_text=True)

    if not any_bug:
        r += 1
        ws.cell(r, 1, "（此解析度未發現 BUG）")
